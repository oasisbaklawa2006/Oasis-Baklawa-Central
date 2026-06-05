#!/usr/bin/env python3
"""Inspect customer master audit workbook and validate against known schema.

Primary file (optional on disk):
  /mnt/data/oasis_customer_master_audit_import_plan.xlsx

When the file is missing, prints the known provisional schema from
scripts/customer_master_workbook_schema.json and exits 0.
Does NOT import into database.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment, misc]

AUDIT_PATH = Path("/mnt/data/oasis_customer_master_audit_import_plan.xlsx")
SCHEMA_PATH = Path(__file__).resolve().parent / "customer_master_workbook_schema.json"
SAMPLE_ROWS = 2


def load_schema() -> dict:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))


def inspect_workbook(path: Path, schema: dict) -> dict:
    if openpyxl is None:
        return {"path": str(path), "error": "openpyxl_not_installed"}

    if not path.is_file():
        return {"path": str(path), "error": "file_not_found"}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    expected_tabs = schema.get("tabs", {})
    sheets: dict[str, dict] = {}
    tab_validation: dict[str, dict] = {}

    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        header = [None if c is None else str(c).strip() for c in header_row] if header_row else []

        data_count = 0
        samples: list[list] = []
        for row in rows_iter:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            data_count += 1
            if len(samples) < SAMPLE_ROWS:
                samples.append([None if c is None else str(c) for c in row])

        sheets[name] = {
            "header": header,
            "data_row_count": data_count,
            "sample_rows": samples,
        }

        tab_spec = expected_tabs.get(name)
        if tab_spec:
            expected_cols = tab_spec.get("expected_columns", [])
            missing = [col for col in expected_cols if col not in header]
            extra = [col for col in header if col and col not in expected_cols]
            tab_validation[name] = {
                "expected_columns": expected_cols,
                "missing_columns": missing,
                "unexpected_columns": extra,
                "valid": len(missing) == 0,
            }

    wb.close()

    missing_tabs = [tab for tab in expected_tabs if tab not in sheets]
    return {
        "path": str(path),
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
        "missing_tabs": missing_tabs,
        "tab_validation": tab_validation,
        "all_tabs_valid": all(v.get("valid", True) for v in tab_validation.values())
        and len(missing_tabs) == 0,
    }


def main() -> int:
    schema = load_schema()
    report = {
        "known_schema": schema,
        "audit_workbook": inspect_workbook(AUDIT_PATH, schema),
    }

    print(json.dumps(report, indent=2))

    audit = report["audit_workbook"]
    if audit.get("error") == "file_not_found":
        print(
            "\nWorkbook file not found; using known provisional headers from "
            "scripts/customer_master_workbook_schema.json",
            file=sys.stderr,
        )
        return 0

    if audit.get("error"):
        print(f"\nInspection error: {audit['error']}", file=sys.stderr)
        return 1

    if not audit.get("all_tabs_valid", False):
        print("\nHeader validation failed for one or more tabs.", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
