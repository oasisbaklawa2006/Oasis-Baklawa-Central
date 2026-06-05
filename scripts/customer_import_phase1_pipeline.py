#!/usr/bin/env python3
"""Customer Master Phase 1 — raw load + validation only (no promotion).

Reads Tally export workbooks from data/customer-import/, writes evidence reports,
optionally loads customer_import_* staging tables via generated SQL.

Does NOT write to companies, whatsapp_contacts, delivery_addresses, users, or orders.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

try:
    import xlrd
except ImportError:
    xlrd = None  # type: ignore

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = Path(__file__).resolve().parent / "customer_import_phase1_config.json"
SQL_OUT = Path(__file__).resolve().parent / "output" / "phase1_staging_load.sql"


def load_config() -> dict:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def cell_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def normalize_gst(raw: str | None) -> str | None:
    if not raw:
        return None
    cleaned = re.sub(r"[^0-9A-Z]", "", raw.upper())
    return cleaned or None


def normalize_phone_last10(raw: str | None) -> str | None:
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if len(digits) >= 10:
        return digits[-10:]
    return None


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (list, dict)):
        return "'" + json.dumps(value).replace("'", "''") + "'::jsonb"
    return "'" + str(value).replace("'", "''") + "'"


def find_column(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [norm_header(h) for h in headers]
    for alias in aliases:
        alias_n = norm_header(alias)
        for idx, header in enumerate(normalized):
            if header == alias_n or alias_n in header:
                return idx
    return None


@dataclass
class SheetInfo:
    name: str
    header_row: int
    headers: list[str]
    data_rows: list[dict[str, Any]]
    blank_rows: int = 0
    malformed_rows: int = 0
    duplicate_headers: list[str] = field(default_factory=list)


@dataclass
class WorkbookInfo:
    path: Path
    sheets: list[SheetInfo] = field(default_factory=list)

    @property
    def size_bytes(self) -> int:
        return self.path.stat().st_size


def read_xlsx(path: Path) -> list[SheetInfo]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for .xlsx files")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: list[SheetInfo] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        header_row_idx = 0
        headers: list[str] = []
        for i, row in enumerate(rows[:15]):
            cells = [cell_str(c) for c in row]
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 2:
                header_row_idx = i
                headers = [c or f"col_{j}" for j, c in enumerate(cells)]
                break
        if not headers:
            sheets.append(SheetInfo(name=name, header_row=1, headers=[], data_rows=[]))
            continue
        header_counts = Counter(norm_header(h) for h in headers if norm_header(h))
        duplicate_headers = [h for h, c in header_counts.items() if c > 1]
        data_rows: list[dict[str, Any]] = []
        blank_rows = 0
        malformed_rows = 0
        for row in rows[header_row_idx + 1 :]:
            if not any(cell_str(c) for c in row):
                blank_rows += 1
                continue
            record: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                key = header or f"col_{idx}"
                record[key] = row[idx] if idx < len(row) else None
            if len(record) < 2:
                malformed_rows += 1
                continue
            data_rows.append(record)
        sheets.append(
            SheetInfo(
                name=name,
                header_row=header_row_idx + 1,
                headers=headers,
                data_rows=data_rows,
                blank_rows=blank_rows,
                malformed_rows=malformed_rows,
                duplicate_headers=duplicate_headers,
            )
        )
    wb.close()
    return sheets


def read_xls(path: Path) -> list[SheetInfo]:
    if xlrd is None:
        raise RuntimeError("xlrd is required for .xls files")
    book = xlrd.open_workbook(path)
    sheets: list[SheetInfo] = []
    for sheet in book.sheets():
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        header_row_idx = 0
        headers: list[str] = []
        for i, row in enumerate(rows[:15]):
            cells = [cell_str(c) for c in row]
            if len([c for c in cells if c]) >= 2:
                header_row_idx = i
                headers = [c or f"col_{j}" for j, c in enumerate(cells)]
                break
        data_rows: list[dict[str, Any]] = []
        blank_rows = 0
        malformed_rows = 0
        header_counts = Counter(norm_header(h) for h in headers if norm_header(h))
        duplicate_headers = [h for h, c in header_counts.items() if c > 1]
        for row in rows[header_row_idx + 1 :]:
            if not any(cell_str(c) for c in row):
                blank_rows += 1
                continue
            record = {headers[i] if i < len(headers) else f"col_{i}": row[i] if i < len(row) else None for i in range(len(headers))}
            if len(record) < 2:
                malformed_rows += 1
                continue
            data_rows.append(record)
        sheets.append(
            SheetInfo(
                name=sheet.name,
                header_row=header_row_idx + 1,
                headers=headers,
                data_rows=data_rows,
                blank_rows=blank_rows,
                malformed_rows=malformed_rows,
                duplicate_headers=duplicate_headers,
            )
        )
    return sheets


def read_workbook(path: Path) -> WorkbookInfo:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        sheets = read_xlsx(path)
    elif suffix == ".xls":
        sheets = read_xls(path)
    else:
        raise ValueError(f"Unsupported workbook type: {path}")
    return WorkbookInfo(path=path, sheets=sheets)


def discover_files(data_dir: Path, expected: list[str]) -> list[Path]:
    found: list[Path] = []
    if not data_dir.is_dir():
        return found
    names = {p.name.lower(): p for p in data_dir.iterdir() if p.is_file()}
    for name in expected:
        p = names.get(name.lower())
        if p:
            found.append(p)
    for p in sorted(data_dir.iterdir()):
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"} and p not in found:
            found.append(p)
    return found


def map_row(headers: list[str], row: dict[str, Any], aliases: dict[str, list[str]]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for field_name, field_aliases in aliases.items():
        idx = find_column(headers, field_aliases)
        if idx is None:
            continue
        header = headers[idx]
        mapped[field_name] = row.get(header)
    return mapped


def build_party_candidates(
    wb: WorkbookInfo, aliases: dict[str, list[str]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    companies: list[dict[str, Any]] = []
    contacts: list[dict[str, Any]] = []
    for sheet in wb.sheets:
        for row_num, row in enumerate(sheet.data_rows, start=sheet.header_row + 1):
            mapped = map_row(sheet.headers, row, aliases)
            name = cell_str(mapped.get("business_name")) or cell_str(row.get(sheet.headers[0]))
            if not name:
                continue
            source_key = f"{sheet.name}:{row_num}"
            gst_raw = cell_str(mapped.get("gst"))
            phone_raw = cell_str(mapped.get("phone_primary"))
            phone_secondary = cell_str(mapped.get("phone_secondary"))
            companies.append(
                {
                    "source_customer_key": source_key,
                    "source_sheet": sheet.name,
                    "source_row": row_num,
                    "business_name": name,
                    "gst_number_raw": gst_raw,
                    "gst_number_normalized": normalize_gst(gst_raw),
                    "registered_address": cell_str(mapped.get("registered_address")),
                    "state": cell_str(mapped.get("state")),
                    "country": cell_str(mapped.get("country")),
                    "phone_raw": phone_raw,
                    "phone_secondary_raw": phone_secondary,
                    "phone_last10": normalize_phone_last10(phone_raw),
                    "phone_secondary_last10": normalize_phone_last10(phone_secondary),
                    "account_manager_email_raw": cell_str(mapped.get("email")),
                    "contact_name": cell_str(mapped.get("contact_name")),
                }
            )
            phone_candidates = [phone_raw, phone_secondary]
            for phone_idx, phone in enumerate(phone_candidates):
                if not phone:
                    continue
                contacts.append(
                    {
                        "source_contact_key": f"{source_key}:phone{phone_idx + 1}",
                        "source_customer_key": source_key,
                        "company_name": name,
                        "source_sheet": sheet.name,
                        "source_row": row_num,
                        "contact_name": cell_str(mapped.get("contact_name")) or name,
                        "whatsapp_phone_raw": phone,
                        "phone_last10": normalize_phone_last10(phone),
                    }
                )
    return companies, contacts


def build_sale_intelligence(wb: WorkbookInfo, aliases: dict[str, list[str]]) -> dict[str, dict[str, Any]]:
    intel: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"invoice_count": 0, "revenue": 0.0, "first_sale_date": None, "last_sale_date": None}
    )
    for sheet in wb.sheets:
        party_idx = find_column(sheet.headers, aliases["business_name"] + ["party", "customer"])
        date_idx = find_column(sheet.headers, aliases["sale_date"])
        amount_idx = find_column(sheet.headers, aliases["sale_amount"])
        if party_idx is None:
            continue
        for row in sheet.data_rows:
            party = cell_str(row.get(sheet.headers[party_idx]))
            if not party:
                continue
            amount = 0.0
            if amount_idx is not None:
                raw_amount = row.get(sheet.headers[amount_idx])
                try:
                    amount = float(str(raw_amount).replace(",", "")) if raw_amount not in (None, "") else 0.0
                except ValueError:
                    amount = 0.0
            entry = intel[party]
            entry["invoice_count"] += 1
            entry["revenue"] += amount
            if date_idx is not None:
                raw_date = row.get(sheet.headers[date_idx])
                if raw_date:
                    ds = str(raw_date)[:10]
                    if not entry["first_sale_date"] or ds < entry["first_sale_date"]:
                        entry["first_sale_date"] = ds
                    if not entry["last_sale_date"] or ds > entry["last_sale_date"]:
                        entry["last_sale_date"] = ds
    return dict(intel)


def build_outstanding_map(wb: WorkbookInfo, aliases: dict[str, list[str]]) -> dict[str, float]:
    outstanding: dict[str, float] = {}
    for sheet in wb.sheets:
        party_idx = find_column(sheet.headers, aliases["business_name"] + ["party", "customer", "ledger"])
        amount_idx = find_column(sheet.headers, aliases["outstanding_amount"])
        if party_idx is None or amount_idx is None:
            continue
        for row in sheet.data_rows:
            party = cell_str(row.get(sheet.headers[party_idx]))
            if not party:
                continue
            raw_amount = row.get(sheet.headers[amount_idx])
            try:
                outstanding[party] = float(str(raw_amount).replace(",", "")) if raw_amount not in (None, "") else 0.0
            except ValueError:
                outstanding[party] = 0.0
    return outstanding


def build_address_records(wb: WorkbookInfo, aliases: dict[str, list[str]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for sheet in wb.sheets:
        party_idx = find_column(sheet.headers, aliases["business_name"] + ["party", "customer", "ledger"])
        addr_idx = find_column(sheet.headers, aliases["address_line"])
        if party_idx is None:
            continue
        city_idx = find_column(sheet.headers, aliases["city"])
        pin_idx = find_column(sheet.headers, aliases["pincode"])
        for row_num, row in enumerate(sheet.data_rows, start=sheet.header_row + 1):
            party = cell_str(row.get(sheet.headers[party_idx]))
            if not party:
                continue
            address = cell_str(row.get(sheet.headers[addr_idx])) if addr_idx is not None else None
            records.append(
                {
                    "company_name": party,
                    "address": address,
                    "city": cell_str(row.get(sheet.headers[city_idx])) if city_idx is not None else None,
                    "pincode": cell_str(row.get(sheet.headers[pin_idx])) if pin_idx is not None else None,
                    "source_sheet": sheet.name,
                    "source_row": row_num,
                }
            )
    return records


def compute_duplicate_review(companies: list[dict[str, Any]], contacts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reviews: list[dict[str, Any]] = []

    gst_groups: dict[str, list[str]] = defaultdict(list)
    for c in companies:
        gst = c.get("gst_number_normalized")
        if gst:
            gst_groups[gst].append(c["source_customer_key"])
    for gst, keys in gst_groups.items():
        if len(keys) > 1:
            reviews.append(
                {
                    "duplicate_type": "gst",
                    "duplicate_key_raw": gst,
                    "duplicate_key_normalized": gst,
                    "occurrence_count": len(keys),
                    "candidate_source_keys": keys,
                }
            )

    name_groups: dict[str, list[str]] = defaultdict(list)
    for c in companies:
        name = norm_header(c.get("business_name"))
        if name:
            name_groups[name].append(c["source_customer_key"])
    for name, keys in name_groups.items():
        if len(keys) > 1:
            reviews.append(
                {
                    "duplicate_type": "name",
                    "duplicate_key_raw": name,
                    "duplicate_key_normalized": name,
                    "occurrence_count": len(keys),
                    "candidate_source_keys": keys,
                }
            )

    phone_groups: dict[str, list[str]] = defaultdict(list)
    for c in companies:
        for phone in [c.get("phone_last10"), c.get("phone_secondary_last10")]:
            if phone:
                phone_groups[phone].append(c["source_customer_key"])
    for ct in contacts:
        phone = ct.get("phone_last10")
        if phone:
            phone_groups[phone].append(ct["source_contact_key"])
    for phone, keys in phone_groups.items():
        if len(set(keys)) > 1:
            reviews.append(
                {
                    "duplicate_type": "phone",
                    "duplicate_key_raw": phone,
                    "duplicate_key_normalized": phone,
                    "occurrence_count": len(keys),
                    "candidate_source_keys": keys,
                }
            )
    return reviews


def generate_load_sql(workbooks: dict[str, WorkbookInfo], companies: list[dict], contacts: list[dict], reviews: list[dict]) -> str:
    lines = [
        "-- Customer Master Phase 1 staging load (generated)",
        "-- staging only — no promotion",
        "BEGIN;",
    ]
    batch_ids: dict[str, str] = {}
    for label, wb in workbooks.items():
        batch_id = hashlib.md5(f"phase1:{label}:{wb.path.name}".encode()).hexdigest()
        batch_uuid = f"{batch_id[:8]}-{batch_id[8:12]}-{batch_id[12:16]}-{batch_id[16:20]}-{batch_id[20:32]}"
        batch_ids[label] = batch_uuid
        row_counts = {s.name: len(s.data_rows) for s in wb.sheets}
        lines.append(
            "INSERT INTO public.customer_import_batches (id, source_filename, source_original_filename, status, row_counts, notes)"
            f" VALUES ({sql_literal(batch_uuid)}, {sql_literal(wb.path.name)}, {sql_literal(wb.path.name)}, 'loaded',"
            f" {sql_literal(row_counts)}::jsonb, 'phase1 raw load')"
            " ON CONFLICT DO NOTHING;"
        )
        for sheet in wb.sheets:
            for row_num, row in enumerate(sheet.data_rows, start=sheet.header_row + 1):
                row_json = json.dumps(row, default=str)
                row_hash = hashlib.sha256(f"{batch_uuid}:{sheet.name}:{row_num}:{row_json}".encode()).hexdigest()
                lines.append(
                    "INSERT INTO public.customer_import_raw (batch_id, source_tab, source_row_number, row_json, row_hash)"
                    f" VALUES ({sql_literal(batch_uuid)}, {sql_literal(sheet.name)}, {row_num},"
                    f" {sql_literal(row_json)}::jsonb, {sql_literal(row_hash)})"
                    " ON CONFLICT DO NOTHING;"
                )

    party_batch = batch_ids.get("party")
    if party_batch and companies:
        for c in companies:
            lines.append(
                "INSERT INTO public.customer_import_company_candidates ("
                "batch_id, source_customer_key, source_sheet, source_row, business_name,"
                "gst_number_raw, gst_number_normalized, registered_address, state, country,"
                "phone_raw, phone_secondary_raw, phone_last10, phone_secondary_last10, account_manager_email_raw)"
                f" VALUES ({sql_literal(party_batch)}, {sql_literal(c['source_customer_key'])},"
                f" {sql_literal(c.get('source_sheet'))}, {sql_literal(c.get('source_row'))},"
                f" {sql_literal(c['business_name'])}, {sql_literal(c.get('gst_number_raw'))},"
                f" {sql_literal(c.get('gst_number_normalized'))}, {sql_literal(c.get('registered_address'))},"
                f" {sql_literal(c.get('state'))}, {sql_literal(c.get('country'))},"
                f" {sql_literal(c.get('phone_raw'))}, {sql_literal(c.get('phone_secondary_raw'))},"
                f" {sql_literal(c.get('phone_last10'))}, {sql_literal(c.get('phone_secondary_last10'))},"
                f" {sql_literal(c.get('account_manager_email_raw'))})"
                " ON CONFLICT (batch_id, source_customer_key) DO NOTHING;"
            )
        for ct in contacts:
            lines.append(
                "INSERT INTO public.customer_import_contact_candidates ("
                "batch_id, source_contact_key, source_customer_key, company_name, source_sheet, source_row,"
                "contact_name, whatsapp_phone_raw, phone_last10)"
                f" VALUES ({sql_literal(party_batch)}, {sql_literal(ct['source_contact_key'])},"
                f" {sql_literal(ct['source_customer_key'])}, {sql_literal(ct['company_name'])},"
                f" {sql_literal(ct.get('source_sheet'))}, {sql_literal(ct.get('source_row'))},"
                f" {sql_literal(ct.get('contact_name'))}, {sql_literal(ct['whatsapp_phone_raw'])},"
                f" {sql_literal(ct.get('phone_last10'))})"
                " ON CONFLICT (batch_id, source_contact_key) DO NOTHING;"
            )
        for review in reviews:
            lines.append(
                "INSERT INTO public.customer_import_duplicate_review ("
                "batch_id, duplicate_type, duplicate_key_raw, duplicate_key_normalized, occurrence_count, candidate_source_keys)"
                f" VALUES ({sql_literal(party_batch)}, {sql_literal(review['duplicate_type'])},"
                f" {sql_literal(review['duplicate_key_raw'])}, {sql_literal(review['duplicate_key_normalized'])},"
                f" {review['occurrence_count']}, ARRAY[{', '.join(sql_literal(k) for k in review['candidate_source_keys'])}])"
                " ON CONFLICT (batch_id, duplicate_type, duplicate_key_normalized) DO NOTHING;"
            )
    lines.append("COMMIT;")
    return "\n".join(lines)


def write_file_inventory(path: Path, files: list[Path], workbooks: dict[str, WorkbookInfo]) -> None:
    lines = [
        "# Customer Import — File Inventory",
        "",
        f"Generated: {datetime.utcnow().isoformat()}Z",
        "",
        "| File | Size (bytes) | Sheets | Total data rows | Column counts |",
        "|------|--------------|--------|-----------------|---------------|",
    ]
    if not files:
        lines.append("| _none found_ | — | — | — | — |")
    for fp in files:
        wb = workbooks.get(fp.name)
        if not wb:
            lines.append(f"| {fp.name} | {fp.stat().st_size} | _unreadable_ | — | — |")
            continue
        total_rows = sum(len(s.data_rows) for s in wb.sheets)
        col_counts = ", ".join(f"{s.name}: {len(s.headers)}" for s in wb.sheets)
        lines.append(
            f"| {fp.name} | {wb.size_bytes} | {len(wb.sheets)} | {total_rows} | {col_counts} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_workbook_structure(path: Path, workbooks: dict[str, WorkbookInfo]) -> None:
    lines = ["# Customer Import — Workbook Structure", "", f"Generated: {datetime.utcnow().isoformat()}Z", ""]
    for name, wb in workbooks.items():
        lines.append(f"## {name}")
        lines.append("")
        for sheet in wb.sheets:
            lines.append(f"### Sheet: {sheet.name}")
            lines.append(f"- Header row: {sheet.header_row}")
            lines.append(f"- Headers: `{', '.join(sheet.headers)}`")
            lines.append(f"- Data rows: {len(sheet.data_rows)}")
            lines.append(f"- Blank rows skipped: {sheet.blank_rows}")
            lines.append(f"- Malformed rows skipped: {sheet.malformed_rows}")
            lines.append(f"- Duplicate headers: {sheet.duplicate_headers or 'none'}")
            lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_customer_intelligence(
    path: Path,
    companies: list[dict[str, Any]],
    sale_intel: dict[str, dict[str, Any]],
    outstanding: dict[str, float],
) -> None:
    lines = ["# Customer Import — Customer Intelligence", "", f"Generated: {datetime.utcnow().isoformat()}Z", ""]
    lines.append("| Company | First sale | Last sale | Invoices | Revenue | Outstanding |")
    lines.append("|---------|------------|-----------|----------|---------|-------------|")
    names = {c["business_name"] for c in companies}
    names.update(sale_intel.keys())
    names.update(outstanding.keys())
    for name in sorted(names):
        s = sale_intel.get(name, {})
        lines.append(
            f"| {name} | {s.get('first_sale_date') or '—'} | {s.get('last_sale_date') or '—'} |"
            f" {s.get('invoice_count', 0)} | {s.get('revenue', 0):.2f} | {outstanding.get(name, 0):.2f} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_phone_analysis(path: Path, companies: list[dict], contacts: list[dict], reviews: list[dict]) -> None:
    lines = ["# Customer Import — Phone Analysis", "", f"Generated: {datetime.utcnow().isoformat()}Z", ""]
    phone_reviews = [r for r in reviews if r["duplicate_type"] == "phone"]
    lines.append(f"- Duplicate phone groups: **{len(phone_reviews)}**")
    lines.append(f"- Company rows with phone: **{sum(1 for c in companies if c.get('phone_last10'))}**")
    lines.append(f"- Contact candidates: **{len(contacts)}**")
    lines.append("")
    for review in phone_reviews[:50]:
        lines.append(f"- `{review['duplicate_key_normalized']}` → {review['occurrence_count']} occurrences")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_gst_analysis(path: Path, companies: list[dict], reviews: list[dict]) -> None:
    lines = ["# Customer Import — GST Analysis", "", f"Generated: {datetime.utcnow().isoformat()}Z", ""]
    with_gst = [c for c in companies if c.get("gst_number_normalized")]
    missing = [c for c in companies if not c.get("gst_number_raw")]
    invalid = [
        c
        for c in companies
        if c.get("gst_number_raw")
        and (not c.get("gst_number_normalized") or len(c["gst_number_normalized"]) != 15)
    ]
    gst_reviews = [r for r in reviews if r["duplicate_type"] == "gst"]
    lines.append(f"- Companies with normalized GST: **{len(with_gst)}**")
    lines.append(f"- Missing GST: **{len(missing)}**")
    lines.append(f"- Invalid GST format: **{len(invalid)}**")
    lines.append(f"- Duplicate GST groups: **{len(gst_reviews)}**")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_address_analysis(path: Path, companies: list[dict], addresses: list[dict]) -> None:
    lines = ["# Customer Import — Address Analysis", "", f"Generated: {datetime.utcnow().isoformat()}Z", ""]
    company_names = {c["business_name"] for c in companies}
    with_address = {a["company_name"] for a in addresses if a.get("address")}
    missing = company_names - with_address
    coverage = (len(with_address) / len(company_names) * 100) if company_names else 0
    addr_counter = Counter(norm_header(a.get("address")) for a in addresses if a.get("address"))
    dup_addresses = [a for a, c in addr_counter.items() if c > 1 and a]
    lines.append(f"- Address candidate rows: **{len(addresses)}**")
    lines.append(f"- Companies with addresses: **{len(with_address)}**")
    lines.append(f"- Companies missing addresses: **{len(missing)}**")
    lines.append(f"- Address coverage: **{coverage:.1f}%**")
    lines.append(f"- Duplicate normalized addresses: **{len(dup_addresses)}**")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_master_match_placeholder(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "# Customer Import — Master Match Analysis",
                "",
                f"Generated: {datetime.utcnow().isoformat()}Z",
                "",
                "Read-only match against `companies`, `shadow_clients`, and `whatsapp_contacts`",
                "runs after staging load via SQL views:",
                "",
                "- `v_customer_import_gst_match_existing`",
                "- `v_customer_import_phone_match_existing`",
                "",
                "Re-run pipeline with `--match-report` after SQL load is applied to populate counts.",
                "",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def write_validation_placeholder(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "# Customer Import — Validation Results",
                "",
                f"Generated: {datetime.utcnow().isoformat()}Z",
                "",
                "Run after staging load:",
                "",
                "```sql",
                "SELECT * FROM public.run_customer_import_validation('<party_batch_id>');",
                "```",
                "",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def write_readiness(
    path: Path,
    companies: list[dict],
    contacts: list[dict],
    addresses: list[dict],
    reviews: list[dict],
) -> None:
    gst_dupes = sum(1 for r in reviews if r["duplicate_type"] == "gst")
    phone_dupes = sum(1 for r in reviews if r["duplicate_type"] == "phone")
    blocked = len(companies) + len(contacts)
    lines = [
        "# Customer Import — Readiness Score",
        "",
        f"Generated: {datetime.utcnow().isoformat()}Z",
        "",
        "| Metric | Count |",
        "|--------|------:|",
        f"| Total companies (candidates) | {len(companies)} |",
        f"| Total contacts | {len(contacts)} |",
        f"| Total phones (contact + company slots) | {sum(1 for c in companies if c.get('phone_last10')) + sum(1 for c in companies if c.get('phone_secondary_last10')) + len(contacts)} |",
        f"| Total GST present | {sum(1 for c in companies if c.get('gst_number_normalized'))} |",
        f"| Total address candidates | {len(addresses)} |",
        f"| Duplicate GST groups | {gst_dupes} |",
        f"| Duplicate phone groups | {phone_dupes} |",
        f"| Rows blocked (default import_action=review) | {blocked} |",
        f"| Rows safe for promotion | 0 |",
        "",
        "**Readiness %:** 0% (promotion blocked until validation gates pass and duplicate review resolved)",
        "",
        "**Promotion remains blocked:** YES",
        "",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_blocked_evidence(evidence_dir: Path, data_dir: Path, expected: list[str]) -> None:
    evidence_dir.mkdir(parents=True, exist_ok=True)
    msg = (
        f"BLOCKED: No workbooks found in `{data_dir}`. "
        f"Place {', '.join(expected)} in that directory and re-run."
    )
    for name, title in [
        ("01_file_inventory.md", "File Inventory"),
        ("02_workbook_structure.md", "Workbook Structure"),
        ("03_customer_intelligence.md", "Customer Intelligence"),
        ("04_phone_analysis.md", "Phone Analysis"),
        ("05_gst_analysis.md", "GST Analysis"),
        ("06_address_analysis.md", "Address Analysis"),
        ("07_master_match_analysis.md", "Master Match Analysis"),
        ("08_validation_results.md", "Validation Results"),
        ("09_import_readiness.md", "Import Readiness"),
    ]:
        (evidence_dir / name).write_text(
            f"# Customer Import — {title}\n\n"
            f"Generated: {datetime.utcnow().isoformat()}Z\n\n"
            f"**Status:** {msg}\n",
            encoding="utf-8",
        )


def main() -> int:
    parser = argparse.ArgumentParser(description="Customer Master Phase 1 pipeline")
    parser.add_argument("--emit-sql", action="store_true", help="Write staging load SQL to scripts/output/")
    parser.add_argument("--data-dir", type=Path, default=None)
    args = parser.parse_args()

    cfg = load_config()
    data_dir = args.data_dir or (ROOT / cfg["data_dir"])
    evidence_dir = ROOT / cfg["evidence_dir"]
    expected = cfg["expected_files"]
    aliases = cfg["column_aliases"]

    files = discover_files(data_dir, expected)
    if not files:
        write_blocked_evidence(evidence_dir, data_dir, expected)
        print(msg := f"BLOCKED: no files in {data_dir}", file=sys.stderr)
        return 2

    workbooks: dict[str, WorkbookInfo] = {}
    for fp in files:
        try:
            workbooks[fp.name] = read_workbook(fp)
        except Exception as exc:  # noqa: BLE001
            print(f"Failed to read {fp}: {exc}", file=sys.stderr)
            return 1

    evidence_dir.mkdir(parents=True, exist_ok=True)
    write_file_inventory(evidence_dir / "01_file_inventory.md", files, workbooks)
    write_workbook_structure(evidence_dir / "02_workbook_structure.md", workbooks)

    companies: list[dict[str, Any]] = []
    contacts: list[dict[str, Any]] = []
    sale_intel: dict[str, dict[str, Any]] = {}
    outstanding: dict[str, float] = {}
    addresses: list[dict[str, Any]] = []

    party_wb = workbooks.get(cfg["party_file"])
    if party_wb:
        companies, contacts = build_party_candidates(party_wb, aliases)
    if cfg["sale_file"] in workbooks:
        sale_intel = build_sale_intelligence(workbooks[cfg["sale_file"]], aliases)
    if cfg["outstanding_file"] in workbooks:
        outstanding = build_outstanding_map(workbooks[cfg["outstanding_file"]], aliases)
    if cfg["address_file"] in workbooks:
        addresses = build_address_records(workbooks[cfg["address_file"]], aliases)

    reviews = compute_duplicate_review(companies, contacts)

    write_customer_intelligence(evidence_dir / "03_customer_intelligence.md", companies, sale_intel, outstanding)
    write_phone_analysis(evidence_dir / "04_phone_analysis.md", companies, contacts, reviews)
    write_gst_analysis(evidence_dir / "05_gst_analysis.md", companies, reviews)
    write_address_analysis(evidence_dir / "06_address_analysis.md", companies, addresses)
    write_master_match_placeholder(evidence_dir / "07_master_match_analysis.md")
    write_validation_placeholder(evidence_dir / "08_validation_results.md")
    write_readiness(evidence_dir / "09_import_readiness.md", companies, contacts, addresses, reviews)

    summary = {
        "files_loaded": [f.name for f in files],
        "raw_rows": sum(sum(len(s.data_rows) for s in wb.sheets) for wb in workbooks.values()),
        "company_candidates": len(companies),
        "contact_candidates": len(contacts),
        "address_candidates": len(addresses),
        "duplicate_gst_groups": sum(1 for r in reviews if r["duplicate_type"] == "gst"),
        "duplicate_phone_groups": sum(1 for r in reviews if r["duplicate_type"] == "phone"),
        "revenue_total": sum(v.get("revenue", 0) for v in sale_intel.values()),
        "outstanding_total": sum(outstanding.values()),
        "promotion_blocked": True,
        "readiness_pct": 0,
    }
    (evidence_dir / "phase1_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    if args.emit_sql:
        SQL_OUT.parent.mkdir(parents=True, exist_ok=True)
        SQL_OUT.write_text(generate_load_sql(workbooks, companies, contacts, reviews), encoding="utf-8")
        print(f"Wrote {SQL_OUT}")

    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
